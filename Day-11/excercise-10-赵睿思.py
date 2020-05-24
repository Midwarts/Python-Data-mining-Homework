from openpyxl import Workbook
import json

# 创建一个Workbook对象
wb = Workbook()
ws = wb.active
# 创建一个Sheet对象
sheet1 = wb.create_sheet(index=0, title="sheet1")

#写入表头
result_head = ['类别','发布时间','发布者','弹幕内容']
result_col = ['类别','发布时间','发布者','弹幕内容']
ws['A1'] = '类别'
ws['B1'] = '发布时间'
ws['C1'] = '发布者'
ws['D1'] = '弹幕内容'
#写入数据
with open('C:/Users/86130/Desktop/萝莉酱直播间时间切片弹幕.json', 'r',encoding='utf-8') as f:
    data = f.read()
    json_data = json.loads(data)
    json_data = json_data["data"]
    for i in range(0, len(json_data)):
        for j in range(4):
            ws.cell(i+2,j+1).value=json_data[i][j]
# 最后保存workbook
wb.save("result.xlsx")
print('操作完成!')