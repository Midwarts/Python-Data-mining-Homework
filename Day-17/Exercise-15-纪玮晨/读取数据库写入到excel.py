import pymysql
from openpyxl import Workbook

#连接数据库
db = pymysql.connect("rm-m5e63zf8ii67j7uq1zo.mysql.rds.aliyuncs.com","huabang_learning","huabang_201904","huabang_learning",charset = "utf8")

#使用cursor()方法创建游标对象
cursor = db.cursor()

sql = "select * from `jwc`"
cursor.execute(sql)#读取数据
all_data = cursor.fetchall()#所有数据

fields = [i[0] for i in cursor.description]#读取表头
print(fields)

wb = Workbook()#创建一个workbook
ws = wb.create_sheet("Mysheet")

ws.cell(row = 1, column = 1).value = fields[0]
ws.cell(row = 1, column = 2).value = fields[1]
ws.cell(row = 1, column = 3).value = fields[2]

for i in all_data:
	print(i)

num = 2
for i in all_data:
	ws.cell(row = num, column = 1).value = i[0]
	ws.cell(row = num, column = 2).value = i[1]
	ws.cell(row = num, column = 3).value = i[2]
	num += 1

wb.save("gdp.xlsx")

db.close()