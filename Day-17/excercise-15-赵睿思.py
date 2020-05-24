import win32com.client
import pypyodbc
from openpyxl import Workbook
import json
wb = Workbook()
ws = wb.active
sheet1 = wb.create_sheet(index=0, title="sheet1")
#获取Connection对象
path=r'C:\Users\86130\Documents\GitHub\Python-Data-mining-Homework\Day-16\excercise-14-赵睿思'
conn = win32com.client.Dispatch('ADODB.Connection')
DSN = 'PROVIDER=Microsoft.ACE.OLEDB.12.0;DATA SOURCE=' + path + ';'
conn.Open(DSN)
rs = win32com.client.Dispatch(r'ADODB.Recordset')
rs.Cursorlocation = 3
rs.Open('SELECT TOP 1 * FROM Sheet1',conn)
for i in range(0, rs.Fields.Count):
    print(rs.Fields[i].Name)
conn = pypyodbc.connect(r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + path + ";Uid=;Pwd=;")
cursor = conn.cursor()
SQL = "SELECT * from 'Sheet1';"
cursor.execute(SQL)
data1 = cursor.fetchall()
str = list(data1[0]),list(data1[1]),list(data1[2]),list(data1[3]),list(data1[4])

data1 = list(str)
data1 = data1["data"]
for i in range(0, len(data1)):
    for j in range(5):
        ws.cell(i+1,j+1).value=data1[i][j]
wb.save("result.xlsx")
print('xls操作完成!')

data1 = json.loads(data1)
json_data = json.dumps(data1, indent=5)
with open('test_data.json', 'w') as json_file:
    json_file.write(json_data)
print('json操作完成！')
