import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
import json

list=[]
#连接数据库
conn = pymysql.connect(user='huabang_learning', host='rm-m5e63zf8ii67j7uq1zo.mysql.rds.aliyuncs.com',port=3306, passwd='huabang_201904', db='huabang_learning', charset='utf8')
cursor=conn.cursor()  #获取浮码
cursor.execute("select * from yjk")       #执行读取操作
x=cursor.fetchall()       #获取数据库
conn.close()   #关闭数据库

for i in x:
    list.append(i)
def write_excel():
    wb=load_workbook(r'C:\Users\NealPayne\Desktop\青村有你.xlsx')
    ws=wb.create_sheet('前九')
    row_1 = ['姓名', '年龄', '家乡', '排名']
    ws.append(row_1)
    for a in list:  # sheet.append循环写入excel
        ws.append(a)  # ！这里不能写成ws.append(list) 不然会直接报错！
    wb.save(r'C:\Users\NealPayne\Desktop\青村有你.xlsx')
    print('操作完成！')

write_excel()