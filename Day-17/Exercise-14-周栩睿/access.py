import pypyodbc
import openpyxl
import win32com.client
import json
path=r'D:\\python\\access.accdb'# 数据库文件
con=win32com.client.Dispatch(r'ADODB.Connection')
DSN="PROVIDER=Microsoft.ACE.OLEDB.12.0;DATA SOURCE="+path+";"
con.Open(DSN)
rs=win32com.client.Dispatch(r'ADODB.Recordset')
rs.Cursorlocation=3
rs.Open('SELECT TOP 1 * FROM 表1',con)
for i in range(0,rs.Fields.Count):
    print(rs.Fields[i].Name)
conn = pypyodbc.connect(r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + path + ";Uid=;Pwd=;")
cursor = conn.cursor()
SQL = "SELECT 表1.[编号], 表1.[名字], 表1.[年龄], 表1.[收养时间]FROM 表1;"#
cursor.execute(SQL)
data1 = cursor.fetchall()
s = list(data1[0]),list(data1[1]),list(data1[2]),list(data1[3])
data1 = list(s)
print(data1)
excel = openpyxl.Workbook()
worksheet = excel.create_sheet("mysheet", index=0)
directionCell = worksheet.cell(row=1, column=1)
directionCell.value = "编号"
directionCell = worksheet.cell(row=1, column=2)
directionCell.value = "名字"
directionCell = worksheet.cell(row=1, column=3)
directionCell.value = "年龄"
directionCell = worksheet.cell(row=1, column=4)
directionCell.value = "收养时间"
for i in range(1,5):
    worksheet.cell(row=2, column=i).value = data1[0][i - 1]
    worksheet.cell(row=3, column=i).value = data1[1][i - 1]
    worksheet.cell(row=4, column=i).value = data1[2][i - 1]
    worksheet.cell(row=5, column=i).value = data1[3][i - 1]
excel.save("D:\\python\\test_access.xlsx")
print("写入excel操作完成")